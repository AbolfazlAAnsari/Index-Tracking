#requrements

import sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import time_changer
import math

#Excel
import openpyxl
from pathlib import Path
xmlfile = "Dataset4.xlsx"
wb = openpyxl.load_workbook(xmlfile, data_only=True)
sheets = wb.sheetnames
main_sheet = wb[sheets[0]]
test_sheet = wb[sheets[1]]

max_column_main = main_sheet.max_column
max_column_test = test_sheet.max_column
max_row         = main_sheet.max_row

#Gams
import os
BASE_DIR = os.path.abspath('')

import gams
ws = gams.workspace.GamsWorkspace(working_directory = BASE_DIR)
db = ws.add_database()

i_max = max_row-2
i_python = [str(i) for i in range(1, i_max+1)]

t_max_main = max_column_main-1
t_python_main = [str(t) for t in range(1, t_max_main+1)]

t_max_test = max_column_test-1
t_python_test = [str(t) for t in range(1, t_max_test+1)]


c_python = 10
lower_python = 0.01
upper_python = 0.3

r_python = {}
for row in range(2, i_max+2):
    for col in range(2, t_max_main+2):
        cell = main_sheet.cell(row=row, column=col)
        r_python[(str(row-1), str(col-1))] = cell.value

rtest_python = {}
for row in range(2, i_max+2):
    for col in range(2, t_max_test+2):
        cell = test_sheet.cell(row=row, column=col)
        rtest_python[(str(row-1), str(col-1))] = cell.value


rprime_python = {}
for col in range(2, t_max_main+2):
    cell = main_sheet.cell(row=i_max+2, column=col)
    rprime_python[str(col-1)] = cell.value

rptest_python = {}
for col in range(2, t_max_test+2):
    cell = test_sheet.cell(row=i_max+2, column=col)
    rptest_python[str(col-1)] = cell.value

#Make Database

i_gams = db.add_set("i", 1)
for ip in i_python:
    i_gams.add_record(ip)

t_gams = db.add_set("t", 1)
for ip in t_python_main:
    t_gams.add_record(ip)

c_gams = db.add_parameter("C",0)
c_gams.add_record().value = c_python

lower_gams = db.add_parameter("lower",0)
lower_gams.add_record().value = lower_python

upper_gams = db.add_parameter("upper",0)
upper_gams.add_record().value = upper_python

r_gams = db.add_parameter_dc("r", [i_gams, t_gams])
for ip in i_python:
    for jp in t_python_main:
        r_gams.add_record((ip, jp)).value = r_python[(ip, jp)]

rprime_gams = db.add_parameter_dc("rprime", [t_gams])
for jp in t_python_main:
    rprime_gams.add_record(jp).value = rprime_python[jp]

#########################

opt = ws.add_options()
opt.defines["gdxincname"] = db.name
m = ws.add_job_from_file("index_tracking.gms")

###########################

time_changer.win_back()
m.run(opt, databases = db)
time_changer.win_update()

print("final result to minimum error: ")

for rec in m.out_db["z"]:
    z_python = rec.level
    print(z_python)

for i in m.out_db["stat"]:
    state = str(i).split("=")[-1]
    print("model state is: ",state)

#############################

x = {}
for rec in m.out_db["x"]:
    x[rec.key(0)] = rec.level

delta = {}
for rec in m.out_db["delta"]:
    delta[rec.key(0)] = rec.level

print("######################")
for i in i_python:
    if x[i]>0:
        print("x["+i+"]: "+str(x[i])+", ", end="\n")
print("######################")

#############################

def prod(x):
    t=1
    for i in x:
        t*=i
    return t

x_plot = [int(i)-1 for i in t_python_test]
y1_plot = [prod([1+rptest_python[tp] for tp in t_python_test if int(tp) <= int(t)]) for t in t_python_test]
y2_plot = []


for t in t_python_test:
    sumation = 0
    for i in i_python:
        sumation += prod([1+rtest_python[(i, tp)] for tp in t_python_test if int(tp) <= int(t)])*x[i]
    y2_plot.append(sumation)

plt.plot(x_plot,y1_plot,label = "Real")
plt.plot(x_plot,y2_plot,label = "Predict")
plt.legend()
plt.show()
